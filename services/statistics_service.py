"""
Statistics Service - Advanced Analytics & Visualization
Provides comprehensive statistics, charts, and export functionality
"""
from __future__ import annotations
from typing import Dict, List, Any, Tuple, Optional
from datetime import datetime, timedelta
import re

# Visualization & Export - with fallback handling
try:
    import matplotlib
    matplotlib.use('TkAgg')  # Use Tkinter backend
    import matplotlib.pyplot as plt
    from matplotlib.figure import Figure
    MATPLOTLIB_AVAILABLE = True
except ImportError:
    MATPLOTLIB_AVAILABLE = False
    Figure = None  # type: ignore

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False

try:
    from reportlab.lib.pagesizes import letter, A4
    from reportlab.lib import colors
    from reportlab.lib.units import inch
    from reportlab.platypus import SimpleDocTemplate, Table, TableStyle, Paragraph, Spacer, PageBreak
    from reportlab.lib.styles import getSampleStyleSheet, ParagraphStyle
    REPORTLAB_AVAILABLE = True
except ImportError:
    REPORTLAB_AVAILABLE = False


class StatisticsService:
    """Service for calculating statistics and generating visualizations"""
    
    def __init__(self, db_manager):
        self.db_manager = db_manager
        
        # Vietnamese weekday names
        self.weekday_names = ['CN', 'T2', 'T3', 'T4', 'T5', 'T6', 'T7']
        
        # Event type keywords for classification
        self.event_types = {
            'Họp/Meeting': ['họp', 'meeting', 'gặp', 'thảo luận', 'phỏng vấn'],
            'Khám bệnh': ['khám', 'bác sĩ', 'bệnh viện', 'nha khoa', 'y tế'],
            'Ăn uống': ['ăn', 'cơm', 'trưa', 'tối', 'sáng', 'nhà hàng', 'quán'],
            'Học tập': ['học', 'lớp', 'bài tập', 'thi', 'kiểm tra', 'ôn tập'],
            'Thể thao': ['gym', 'bơi', 'chạy', 'yoga', 'thể dục', 'tennis'],
            'Giải trí': ['phim', 'xem', 'chơi', 'du lịch', 'picnic'],
        }
    
    # ==================== STATISTICS CALCULATION ====================
    
    def get_comprehensive_stats(self) -> Dict[str, Any]:
        """Get all statistics in one call"""
        return {
            'overview': self.get_overview_stats(),
            'time': self.get_time_stats(),
            'location': self.get_location_stats(),
            'event_type': self.get_event_type_stats(),
            'trends': self.get_trend_stats(),
        }
    
    def get_overview_stats(self) -> Dict[str, Any]:
        """Get overview statistics"""
        all_events = self.db_manager.get_all_events()
        total = len(all_events)
        
        now = datetime.now()
        week_start = now - timedelta(days=now.weekday())  # Monday
        month_start = now.replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        
        week_count = sum(1 for e in all_events 
                        if e.get('start_time') and 
                        datetime.fromisoformat(e['start_time'][:19]) >= week_start)
        
        month_count = sum(1 for e in all_events 
                         if e.get('start_time') and 
                         datetime.fromisoformat(e['start_time'][:19]) >= month_start)
        
        # Count events with reminders
        with_reminder = sum(1 for e in all_events if (e.get('reminder_minutes') or 0) > 0)
        
        # Count events with location
        with_location = sum(1 for e in all_events if e.get('location'))
        
        # Calculate streak (consecutive days with events)
        streak = self._calculate_streak(all_events)
        
        # Average events per day (last 30 days)
        thirty_days_ago = now - timedelta(days=30)
        recent_events = [e for e in all_events 
                        if e.get('start_time') and 
                        datetime.fromisoformat(e['start_time'][:19]) >= thirty_days_ago]
        avg_per_day = len(recent_events) / 30.0 if recent_events else 0
        
        return {
            'total_events': total,
            'week_events': week_count,
            'month_events': month_count,
            'with_reminder': with_reminder,
            'with_location': with_location,
            'reminder_percentage': (with_reminder / total * 100) if total > 0 else 0,
            'location_percentage': (with_location / total * 100) if total > 0 else 0,
            'current_streak': streak['current'],
            'longest_streak': streak['longest'],
            'avg_events_per_day': avg_per_day,
        }
    
    def get_time_stats(self) -> Dict[str, Any]:
        """Get time-based statistics"""
        all_events = self.db_manager.get_all_events()
        
        # By weekday (0=Monday, 6=Sunday)
        weekday_counts = [0] * 7
        hour_counts = [0] * 24
        
        for event in all_events:
            if event.get('start_time'):
                try:
                    dt = datetime.fromisoformat(event['start_time'][:19])
                    weekday = dt.weekday()  # 0=Monday
                    hour = dt.hour
                    
                    weekday_counts[weekday] += 1
                    hour_counts[hour] += 1
                except:
                    pass
        
        # Find peak hour and peak day
        peak_hour = hour_counts.index(max(hour_counts)) if max(hour_counts) > 0 else None
        peak_day = weekday_counts.index(max(weekday_counts)) if max(weekday_counts) > 0 else None
        
        return {
            'by_weekday': weekday_counts,
            'by_hour': hour_counts,
            'peak_hour': peak_hour,
            'peak_day': peak_day,
            'peak_hour_count': max(hour_counts) if hour_counts else 0,
            'peak_day_count': max(weekday_counts) if weekday_counts else 0,
        }
    
    def get_location_stats(self) -> Dict[str, Any]:
        """Get location statistics"""
        all_events = self.db_manager.get_all_events()
        
        location_counts = {}
        for event in all_events:
            loc = event.get('location')
            if loc and loc.strip():
                loc = loc.strip()
                location_counts[loc] = location_counts.get(loc, 0) + 1
        
        # Sort by count descending
        sorted_locations = sorted(location_counts.items(), key=lambda x: x[1], reverse=True)
        
        return {
            'top_locations': sorted_locations[:10],  # Top 10
            'total_unique_locations': len(location_counts),
            'total_with_location': sum(location_counts.values()),
        }
    
    def get_event_type_stats(self) -> Dict[str, Any]:
        """Classify and count events by type"""
        all_events = self.db_manager.get_all_events()
        
        type_counts = {k: 0 for k in self.event_types.keys()}
        type_counts['Khác'] = 0
        
        for event in all_events:
            event_name = (event.get('event_name') or '').lower()
            classified = False
            
            for type_name, keywords in self.event_types.items():
                if any(kw in event_name for kw in keywords):
                    type_counts[type_name] += 1
                    classified = True
                    break
            
            if not classified:
                type_counts['Khác'] += 1
        
        # Calculate percentages
        total = sum(type_counts.values())
        type_percentages = {k: (v / total * 100) if total > 0 else 0 
                           for k, v in type_counts.items()}
        
        return {
            'counts': type_counts,
            'percentages': type_percentages,
            'total': total,
        }
    
    def get_trend_stats(self) -> Dict[str, Any]:
        """Get trend analysis for last 30 days"""
        all_events = self.db_manager.get_all_events()
        now = datetime.now()
        
        # Group by week for last 4 weeks
        weekly_counts = [0] * 4
        
        for event in all_events:
            if event.get('start_time'):
                try:
                    dt = datetime.fromisoformat(event['start_time'][:19])
                    days_ago = (now - dt).days
                    
                    if 0 <= days_ago < 7:
                        weekly_counts[0] += 1
                    elif 7 <= days_ago < 14:
                        weekly_counts[1] += 1
                    elif 14 <= days_ago < 21:
                        weekly_counts[2] += 1
                    elif 21 <= days_ago < 28:
                        weekly_counts[3] += 1
                except:
                    pass
        
        # Calculate growth rate
        if weekly_counts[1] > 0:
            growth_rate = ((weekly_counts[0] - weekly_counts[1]) / weekly_counts[1]) * 100
        else:
            growth_rate = 0
        
        return {
            'weekly_counts': list(reversed(weekly_counts)),  # Oldest to newest
            'growth_rate': growth_rate,
        }
    
    def _calculate_streak(self, events: List[Dict]) -> Dict[str, int]:
        """Calculate current and longest streak of days with events"""
        if not events:
            return {'current': 0, 'longest': 0}
        
        # Get all unique dates
        dates = set()
        for event in events:
            if event.get('start_time'):
                try:
                    dt = datetime.fromisoformat(event['start_time'][:19])
                    dates.add(dt.date())
                except:
                    pass
        
        if not dates:
            return {'current': 0, 'longest': 0}
        
        sorted_dates = sorted(dates, reverse=True)
        
        # Calculate current streak
        current_streak = 0
        today = datetime.now().date()
        check_date = today
        
        for i in range(365):  # Max check 1 year
            if check_date in dates:
                current_streak += 1
                check_date -= timedelta(days=1)
            else:
                break
        
        # Calculate longest streak
        longest_streak = 1
        current_run = 1
        
        for i in range(len(sorted_dates) - 1):
            diff = (sorted_dates[i] - sorted_dates[i + 1]).days
            if diff == 1:
                current_run += 1
                longest_streak = max(longest_streak, current_run)
            else:
                current_run = 1
        
        return {'current': current_streak, 'longest': longest_streak}
    
    # ==================== CHART GENERATION ====================
    
    def create_weekday_chart(self, stats: Dict) -> Optional[Any]:
        """Create weekday distribution bar chart"""
        if not MATPLOTLIB_AVAILABLE:
            raise ImportError("matplotlib is required for chart generation")
        
        # Use Figure directly instead of pyplot to avoid GUI conflicts
        fig = Figure(figsize=(8, 5), dpi=100)
        ax = fig.add_subplot(111)
        
        weekdays = ['T2', 'T3', 'T4', 'T5', 'T6', 'T7', 'CN']
        counts = stats['by_weekday']
        
        bars = ax.bar(weekdays, counts, color='skyblue', edgecolor='navy', alpha=0.7)
        
        # Highlight peak day
        if stats['peak_day'] is not None:
            bars[stats['peak_day']].set_color('orange')
        
        ax.set_xlabel('Ngày trong tuần', fontsize=12)
        ax.set_ylabel('Số sự kiện', fontsize=12)
        ax.set_title('Phân bố sự kiện theo ngày trong tuần', fontsize=14, fontweight='bold')
        ax.grid(axis='y', alpha=0.3)
        
        # Add value labels on bars
        for bar in bars:
            height = bar.get_height()
            ax.text(bar.get_x() + bar.get_width()/2., height,
                   f'{int(height)}',
                   ha='center', va='bottom', fontsize=10)
        
        fig.tight_layout()
        return fig
    
    def create_hourly_chart(self, stats: Dict) -> Optional[Any]:
        """Create hourly distribution chart"""
        if not MATPLOTLIB_AVAILABLE:
            raise ImportError("matplotlib is required for chart generation")
        
        fig = Figure(figsize=(10, 5), dpi=100)
        ax = fig.add_subplot(111)
        
        hours = list(range(24))
        counts = stats['by_hour']
        
        bars = ax.bar(hours, counts, color='lightgreen', edgecolor='darkgreen', alpha=0.7)
        
        # Highlight peak hour
        if stats['peak_hour'] is not None:
            bars[stats['peak_hour']].set_color('red')
        
        ax.set_xlabel('Giờ trong ngày', fontsize=12)
        ax.set_ylabel('Số sự kiện', fontsize=12)
        ax.set_title('Phân bố sự kiện theo giờ', fontsize=14, fontweight='bold')
        ax.set_xticks(range(0, 24, 2))
        ax.grid(axis='y', alpha=0.3)
        
        fig.tight_layout()
        return fig
    
    def create_location_chart(self, stats: Dict) -> Optional[Any]:
        """Create top locations bar chart"""
        if not MATPLOTLIB_AVAILABLE:
            raise ImportError("matplotlib is required for chart generation")
        
        fig = Figure(figsize=(8, 6), dpi=100)
        ax = fig.add_subplot(111)
        
        top_locs = stats['top_locations'][:5]  # Top 5
        if not top_locs:
            ax.text(0.5, 0.5, 'Không có dữ liệu địa điểm', 
                   ha='center', va='center', fontsize=14)
            ax.set_xlim(0, 1)
            ax.set_ylim(0, 1)
            ax.axis('off')
        else:
            locations = [loc[0][:20] + '...' if len(loc[0]) > 20 else loc[0] for loc in top_locs]
            counts = [loc[1] for loc in top_locs]
            
            bars = ax.barh(locations, counts, color='coral', edgecolor='darkred', alpha=0.7)
            
            ax.set_xlabel('Số lần', fontsize=12)
            ax.set_title('Top 5 địa điểm thường xuyên', fontsize=14, fontweight='bold')
            ax.grid(axis='x', alpha=0.3)
            
            # Add value labels
            for bar in bars:
                width = bar.get_width()
                ax.text(width, bar.get_y() + bar.get_height()/2.,
                       f' {int(width)}',
                       ha='left', va='center', fontsize=10)
        
        fig.tight_layout()
        return fig
    
    def create_event_type_pie_chart(self, stats: Dict) -> Optional[Any]:
        """Create event type pie chart"""
        if not MATPLOTLIB_AVAILABLE:
            raise ImportError("matplotlib is required for chart generation")
        
        fig = Figure(figsize=(8, 8), dpi=100)
        ax = fig.add_subplot(111)
        
        # Filter out zero counts
        data = {k: v for k, v in stats['counts'].items() if v > 0}
        
        if not data:
            ax.text(0.5, 0.5, 'Không có dữ liệu', 
                   ha='center', va='center', fontsize=14)
            ax.axis('off')
        else:
            labels = list(data.keys())
            sizes = list(data.values())
            
            colors_list = ['#ff9999', '#66b3ff', '#99ff99', '#ffcc99', 
                          '#ff99cc', '#c2c2f0', '#ffb3e6']
            
            wedges, texts, autotexts = ax.pie(
                sizes, 
                labels=labels, 
                autopct='%1.1f%%',
                colors=colors_list[:len(labels)],
                startangle=90,
                textprops={'fontsize': 11}
            )
            
            # Bold percentage text
            for autotext in autotexts:
                autotext.set_color('white')
                autotext.set_fontweight('bold')
            
            ax.set_title('Phân loại sự kiện theo nội dung', fontsize=14, fontweight='bold')
        
        fig.tight_layout()
        return fig
    
    def create_trend_chart(self, stats: Dict) -> Optional[Any]:
        """Create trend line chart"""
        if not MATPLOTLIB_AVAILABLE:
            raise ImportError("matplotlib is required for chart generation")
        
        fig = Figure(figsize=(8, 5), dpi=100)
        ax = fig.add_subplot(111)
        
        weeks = ['Tuần 1', 'Tuần 2', 'Tuần 3', 'Tuần 4']
        counts = stats['weekly_counts']
        
        ax.plot(weeks, counts, marker='o', linewidth=2, markersize=8, 
               color='purple', label='Số sự kiện')
        ax.fill_between(range(len(weeks)), counts, alpha=0.3, color='purple')
        
        ax.set_xlabel('Thời gian', fontsize=12)
        ax.set_ylabel('Số sự kiện', fontsize=12)
        ax.set_title('Xu hướng 4 tuần gần đây', fontsize=14, fontweight='bold')
        ax.grid(True, alpha=0.3)
        ax.legend()
        
        # Add value labels
        for i, count in enumerate(counts):
            ax.text(i, count, f' {count}', ha='left', va='bottom', fontsize=10)
        
        fig.tight_layout()
        return fig
    
    # ==================== EXPORT FUNCTIONS ====================
    
    def export_to_excel(self, filepath: str, stats: Dict) -> None:
        """Export statistics to Excel file"""
        if not OPENPYXL_AVAILABLE:
            raise ImportError("openpyxl is required for Excel export")
        wb = openpyxl.Workbook()
        
        # Remove default sheet
        if 'Sheet' in wb.sheetnames:
            del wb['Sheet']
        
        # Overview sheet
        ws_overview = wb.create_sheet('Tổng quan')
        self._write_overview_sheet(ws_overview, stats['overview'])
        
        # Time sheet
        ws_time = wb.create_sheet('Phân tích thời gian')
        self._write_time_sheet(ws_time, stats['time'])
        
        # Location sheet
        ws_location = wb.create_sheet('Địa điểm')
        self._write_location_sheet(ws_location, stats['location'])
        
        # Event type sheet
        ws_type = wb.create_sheet('Loại sự kiện')
        self._write_event_type_sheet(ws_type, stats['event_type'])
        
        wb.save(filepath)
    
    def _write_overview_sheet(self, ws, stats):
        """Write overview statistics to Excel sheet"""
        # Header
        ws['A1'] = 'THỐNG KÊ TỔNG QUAN'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        row = 3
        data = [
            ('Tổng số sự kiện:', stats['total_events']),
            ('Sự kiện tuần này:', stats['week_events']),
            ('Sự kiện tháng này:', stats['month_events']),
            ('Có nhắc nhở:', f"{stats['with_reminder']} ({stats['reminder_percentage']:.1f}%)"),
            ('Có địa điểm:', f"{stats['with_location']} ({stats['location_percentage']:.1f}%)"),
            ('Streak hiện tại:', f"{stats['current_streak']} ngày"),
            ('Streak dài nhất:', f"{stats['longest_streak']} ngày"),
            ('TB sự kiện/ngày (30 ngày):', f"{stats['avg_events_per_day']:.1f}"),
        ]
        
        for label, value in data:
            ws.cell(row, 1, label).font = Font(bold=True)
            ws.cell(row, 2, value)
            row += 1
        
        ws.column_dimensions['A'].width = 30
        ws.column_dimensions['B'].width = 20
    
    def _write_time_sheet(self, ws, stats):
        """Write time analysis to Excel sheet"""
        ws['A1'] = 'PHÂN TÍCH THỜI GIAN'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        # Weekday distribution
        ws['A3'] = 'Theo ngày trong tuần:'
        ws['A3'].font = Font(bold=True)
        
        weekdays = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ nhật']
        for i, (day, count) in enumerate(zip(weekdays, stats['by_weekday']), start=4):
            ws.cell(i, 1, day)
            ws.cell(i, 2, count)
        
        # Peak info
        row = 12
        ws.cell(row, 1, 'Ngày bận nhất:').font = Font(bold=True)
        if stats['peak_day'] is not None:
            ws.cell(row, 2, f"{weekdays[stats['peak_day']]} ({stats['peak_day_count']} sự kiện)")
        
        row += 1
        ws.cell(row, 1, 'Giờ bận nhất:').font = Font(bold=True)
        if stats['peak_hour'] is not None:
            ws.cell(row, 2, f"{stats['peak_hour']}:00 ({stats['peak_hour_count']} sự kiện)")
        
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 20
    
    def _write_location_sheet(self, ws, stats):
        """Write location statistics to Excel sheet"""
        ws['A1'] = 'TOP ĐỊA ĐIỂM'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:B1')
        
        ws['A3'] = 'Địa điểm'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = 'Số lần'
        ws['B3'].font = Font(bold=True)
        
        for i, (location, count) in enumerate(stats['top_locations'], start=4):
            ws.cell(i, 1, location)
            ws.cell(i, 2, count)
        
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 15
    
    def _write_event_type_sheet(self, ws, stats):
        """Write event type statistics to Excel sheet"""
        ws['A1'] = 'PHÂN LOẠI SỰ KIỆN'
        ws['A1'].font = Font(bold=True, size=14)
        ws.merge_cells('A1:C1')
        
        ws['A3'] = 'Loại'
        ws['A3'].font = Font(bold=True)
        ws['B3'] = 'Số lượng'
        ws['B3'].font = Font(bold=True)
        ws['C3'] = 'Tỷ lệ'
        ws['C3'].font = Font(bold=True)
        
        for i, (type_name, count) in enumerate(stats['counts'].items(), start=4):
            percentage = stats['percentages'][type_name]
            ws.cell(i, 1, type_name)
            ws.cell(i, 2, count)
            ws.cell(i, 3, f"{percentage:.1f}%")
        
        ws.column_dimensions['A'].width = 20
        ws.column_dimensions['B'].width = 15
        ws.column_dimensions['C'].width = 15
    
    def export_to_pdf(self, filepath: str, stats: Dict) -> None:
        """Export statistics to PDF file"""
        if not REPORTLAB_AVAILABLE:
            raise ImportError("reportlab is required for PDF export")
        doc = SimpleDocTemplate(filepath, pagesize=A4)
        story = []
        styles = getSampleStyleSheet()
        
        # Title
        title_style = ParagraphStyle(
            'CustomTitle',
            parent=styles['Heading1'],
            fontSize=18,
            textColor=colors.HexColor('#2E4053'),
            spaceAfter=30,
            alignment=1  # Center
        )
        
        story.append(Paragraph('BÁO CÁO THỐNG KÊ LỊCH TRÌNH', title_style))
        story.append(Spacer(1, 0.2 * inch))
        
        # Overview section
        story.append(Paragraph('1. TỔNG QUAN', styles['Heading2']))
        overview_data = [
            ['Chỉ số', 'Giá trị'],
            ['Tổng số sự kiện', str(stats['overview']['total_events'])],
            ['Sự kiện tuần này', str(stats['overview']['week_events'])],
            ['Sự kiện tháng này', str(stats['overview']['month_events'])],
            ['Có nhắc nhở', f"{stats['overview']['with_reminder']} ({stats['overview']['reminder_percentage']:.1f}%)"],
            ['Có địa điểm', f"{stats['overview']['with_location']} ({stats['overview']['location_percentage']:.1f}%)"],
            ['Streak hiện tại', f"{stats['overview']['current_streak']} ngày"],
            ['Streak dài nhất', f"{stats['overview']['longest_streak']} ngày"],
        ]
        
        overview_table = Table(overview_data, colWidths=[3*inch, 2*inch])
        overview_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.beige),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(overview_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Time analysis section
        story.append(Paragraph('2. PHÂN TÍCH THỜI GIAN', styles['Heading2']))
        weekdays = ['Thứ 2', 'Thứ 3', 'Thứ 4', 'Thứ 5', 'Thứ 6', 'Thứ 7', 'Chủ nhật']
        time_data = [['Ngày', 'Số sự kiện']]
        for day, count in zip(weekdays, stats['time']['by_weekday']):
            time_data.append([day, str(count)])
        
        time_table = Table(time_data, colWidths=[2.5*inch, 2*inch])
        time_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightblue),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(time_table)
        story.append(Spacer(1, 0.3 * inch))
        
        # Location section
        story.append(Paragraph('3. TOP ĐỊA ĐIỂM', styles['Heading2']))
        location_data = [['STT', 'Địa điểm', 'Số lần']]
        for i, (loc, count) in enumerate(stats['location']['top_locations'][:5], 1):
            location_data.append([str(i), loc[:40], str(count)])
        
        if len(location_data) > 1:
            location_table = Table(location_data, colWidths=[0.5*inch, 3*inch, 1*inch])
            location_table.setStyle(TableStyle([
                ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
                ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
                ('ALIGN', (0, 0), (-1, -1), 'LEFT'),
                ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
                ('FONTSIZE', (0, 0), (-1, 0), 12),
                ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
                ('BACKGROUND', (0, 1), (-1, -1), colors.lightgrey),
                ('GRID', (0, 0), (-1, -1), 1, colors.black)
            ]))
            story.append(location_table)
        else:
            story.append(Paragraph('Không có dữ liệu địa điểm', styles['Normal']))
        
        story.append(Spacer(1, 0.3 * inch))
        
        # Event type section
        story.append(Paragraph('4. PHÂN LOẠI SỰ KIỆN', styles['Heading2']))
        type_data = [['Loại', 'Số lượng', 'Tỷ lệ']]
        for type_name, count in stats['event_type']['counts'].items():
            percentage = stats['event_type']['percentages'][type_name]
            type_data.append([type_name, str(count), f"{percentage:.1f}%"])
        
        type_table = Table(type_data, colWidths=[2*inch, 1.5*inch, 1.5*inch])
        type_table.setStyle(TableStyle([
            ('BACKGROUND', (0, 0), (-1, 0), colors.grey),
            ('TEXTCOLOR', (0, 0), (-1, 0), colors.whitesmoke),
            ('ALIGN', (0, 0), (-1, -1), 'CENTER'),
            ('FONTNAME', (0, 0), (-1, 0), 'Helvetica-Bold'),
            ('FONTSIZE', (0, 0), (-1, 0), 12),
            ('BOTTOMPADDING', (0, 0), (-1, 0), 12),
            ('BACKGROUND', (0, 1), (-1, -1), colors.lightgreen),
            ('GRID', (0, 0), (-1, -1), 1, colors.black)
        ]))
        
        story.append(type_table)
        
        # Build PDF
        doc.build(story)
